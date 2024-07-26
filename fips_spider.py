from typing import List, Any
from dotenv import load_dotenv
from selenium.webdriver.firefox.firefox_binary import FirefoxBinary
from seleniumrequests import Firefox
from datetime import datetime, timedelta
from timeit import default_timer as timer
from math import floor
import openpyxl
import os
from pathlib import Path
from program import Program
from filter import Filter

class FipsSpider():
    base_url = 'https://fips.ru/'
    entry_url = 'iiss/db.xhtml'
    search_url = 'iiss/search.xhtml'
    programs: List[Program] = []
    output_col_titles = ['дата регистрации','№ свидетельства', 'тип', 'название', 'авторы', 'правообладатель', 'реферат']
    
    def __init__(self):
        # Load the environment variables from the .env file
        load_dotenv(Path('.') / '.env', verbose=True, override=True)
        self.ids_seen = set()
        self.start_time = timer()
        self.programs_seen = 0
        self.authors_seen = 0
        self.total_authors = 0
        firefox_path = os.environ.get('firefox_binary')
        print(f'{firefox_path=}')
        #assert(firefox_path is not None)
        self.driver = Firefox() if firefox_path is None else Firefox(firefox_binary=FirefoxBinary(firefox_path))
        self.programs_skipped = 0
        self.driver.get(self.base_url + self.entry_url)
        assert(self.driver.title == 'Информационно-поисковая система')
        self.driver.find_element_by_xpath('//form/div[8]').click()
        self.driver.find_element_by_id('db-selection-form:dbsGrid9:0:dbsGrid9checkbox').click()
        self.driver.implicitly_wait(1)
        self.driver.find_element_by_class_name('save').click()
        self.driver.implicitly_wait(3)

    def add_program(self, program):
        self.programs_seen += 1
        if program.id in self.ids_seen:
            pass
            # print(f'пропущена программа {program.id} (уже найдена у другого автора)')
        else:
            self.ids_seen.add(program.id)
            self.programs += [program]

    def parse_program(self):
        res = Program()
        try:
            res.id = self.driver.find_element_by_xpath('//td[@id="top4"]/a').get_attribute("innerHTML")
            res.reg_date = self.driver.find_element_by_xpath('//p[contains(text(),"Дата регистрации:")]/b').get_attribute("innerHTML")
            res.title = self.driver.find_element_by_xpath('//div[@id="mainDoc"]/p[1]/b').get_attribute("innerHTML")
            res.referat = self.driver.find_element_by_xpath('//div[@id="mainDoc"]/p[2]').get_attribute("innerHTML")
            res.authors = self.driver.find_element_by_xpath('//td[@id="bibl"]/p[1]/b').get_attribute("innerHTML")
            res.owner = self.driver.find_element_by_xpath('//td[@id="bibl"]/p[2]/b').get_attribute("innerHTML")
        except Exception as e:
            print(f'ошибка обработки: {e}')
            return None
        # print(f'found {res.id=}')
        # print(f'found {res.authors=}')
        # print(f'found {res.owner=}')
        return res

    def fetch_author_programs(self, credentials: str):
        self.authors_seen += 1
        # переходим на страницу поиска
        self.driver.get(self.base_url + self.search_url)
        # ждем загрузки DOM для очистки поля с автором
        self.driver.implicitly_wait(1)
        # вбиваем имя автора и нажимаем поиск
        search_input = self.driver.find_element_by_xpath('//input[contains(@id,"fields:5")]')
        search_input.clear()
        search_input.send_keys(f'"{credentials}"')
        self.driver.find_element_by_class_name('save').click()
        self.driver.implicitly_wait(1)
        try:
            # открываем ссылку на первую программу
            first_program = self.driver.find_element_by_xpath('//a[@class="tr"]')
            #print(f'found first program {first_program}')
            is_done = False
            self.driver.get(first_program.get_attribute("href"))
        except Exception:
            # print(f'у автора {credentials} программы не найдены')
            is_done = True

        # в цикле пробегаем по всем страницам
        while not is_done:
            program = self.parse_program()
            if program is not None:
                # если правообладатель в белом списке, добавляем программу
                if Filter.is_whitelisted(program.owner):
                    self.add_program(program)
                # иначе, если правообладатель не в черном списке, также добавляем программу
                elif not Filter.is_blacklisted(program.owner):
                    self.add_program(program)
                else:
                    self.programs_skipped += 1
                    #print(f"пропущена программа: п: {program.owner}, a: {program.authors}")
            progress = f'{(self.authors_seen*100/self.total_authors):.2f}'
            speed = ((timer() - self.start_time)/self.authors_seen)
            remains = floor((self.total_authors-self.authors_seen)*speed)
            print(f'[{progress}%][{speed:.1f} с/авт.][осталось ~{timedelta(seconds=remains)}][Программ: {len(self.programs)}][Пропущено: {self.programs_skipped}] Автор #{self.authors_seen} {credentials.split()[0]} (из {self.total_authors}).', end='\r')
            try:
                next_page = self.driver.find_element_by_xpath('//a[@class="ui-link ui-widget modern-page-next"]')
                next_page.click()
            except Exception:
                is_done = True

    def fetch(self, authors):
        self.start_time = timer()
        for author in authors: 
            self.fetch_author_programs(author)

    def write_output(self, file_name: str) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(self.output_col_titles)
        for program in self.programs:
            ws.append([program.reg_date,
                       program.id,
                       program.kind,
                       program.title,
                       program.authors,
                       program.owner,
                       program.referat])
        wb.save(file_name + '.xlsx')
 
    def parse_dir(self, dir_name: str) -> None:
        self.total_authors = 0
        files = []
        files += [fname for fname in os.listdir(dir_name) if fname.endswith('.xlsx')]
        print(f'файлы для обработки: {files}')
        payload = []
        for file in files:
            try:
                wb = openpyxl.load_workbook(Path(dir_name, file))
                ws = wb.active
                if ws is None:
                    continue
                fios = []
                fio_col = next((col for col in ws.iter_cols(1, ws.max_column) if 'фио' in str(col[0].value).lower()), None)
                if fio_col is None:
                    print(f'файл {file} не содержит ФИО в заголовках!')
                    continue      
                fio_col_idx = fio_col[0].col_idx
                print(f'ФИО в стобце {fio_col_idx}')
                for row in ws.iter_rows(2, ws.max_row):
                    fio = row[fio_col_idx - 1].value
                    if fio is None:
                        break
                    fio = str(fio).strip()
                    if fio == '':
                        print(f'некорректное имя в строке {row[fio_col_idx - 1].row}')
                        continue
                    fios += [fio]
                print(f'добавлено {len(fios)} имен из {file}')
                self.total_authors += len(fios)
                payload += [fios]
            except Exception as e:
                print(f'ошибка при обработке файла: {e}')
        return payload

    def process_dir(self, dir_name: str) -> None:
        payload = self.parse_dir(dir_name)
        print(f'К обработке {self.total_authors} авторов.')
        start = timer()
        total = 0
        for names in payload:
            total += len(names)
            self.fetch(names)
        elapsed_time = timer() - start
        now = datetime.now()
        self.write_output(f'Поиск_{now.year}-{now.month}-{now.day}')
        print(f'\nРАБОТА ЗАВЕРШЕНА. Выгружено {len(self.programs)} программ.')
        print(f'Время работы {timedelta(seconds=elapsed_time)}. Производительность: {(elapsed_time/self.total_authors):.2f} сек./автор.')